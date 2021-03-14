from os.path import split
from django.http.response import JsonResponse
from django.shortcuts import render,redirect
from django.http import HttpResponse
from .models import Item,SoldItem,PurchasedItem,ItemCategory,InvoiceSells,PaymentMode
from .forms import InvoiceSellsForm,SoldItemForm
#import pytesseract
from PIL import Image
#from googletrans import Translator
from xc_stock_tracker.settings import BASE_DIR
import os
#import pywhatkit
import datetime
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image
from reportlab.lib.pagesizes import letter,landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet,ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.lib.pagesizes import A3
from reportlab.pdfgen.canvas import Canvas
from django.http import FileResponse
from xc_stock_tracker.settings import STATIC_URL as static
from xc_stock_tracker.settings import BASE_DIR

from selenium import webdriver
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

from time import sleep
import time
import openpyxl as xl

# def readXL():
# 	filename = 'KWALITY PRODUCT SALE REPORT 2021 (1).xlsx'
# 	# filename = 'gradeData.xlsx'
# 	filepath = os.path.join(BASE_DIR,filename)
# 	wb = xl.load_workbook(filepath) #workbook
# 	sheet = wb.active
# 	max_col = sheet.max_column
# 	print('max column :',max_col)
# 	dateStored = None
# 	paymodeStored = None

# 	# invoice = InvoiceSells()
# 	for ri in range(2,22):
# 		# print(ri,end=' ')

# 		product_cell = sheet.cell(row=ri,column=3) #product cell
		
# 		if product_cell.value != None:  
# 			date = sheet.cell(row = ri, column = 1).value #date cell
# 			paymode =sheet.cell(row = ri, column = 6).value #payment mode cell
			
# 			if date != None or paymode != paymodeStored:
# 				if date != None:
# 					dateStored = date
# 				if paymode != paymodeStored:
# 					paymodeStored = paymode

# 				invoice = InvoiceSells()
# 				invoice.dos = dateStored
# 				invoice.discount = 0
# 				invoice.customer = 'Unknown'
# 				invoice.mobile = '9876543210'
# 				invoice.total_amount = 0
# 				invoice.status = True
# 				print(paymode)
# 				print('objs: ',PaymentMode.objects.filter(mode__contains = paymode))
# 				print('objs: ',PaymentMode.objects.filter(mode__contains = paymode).first())
# 				invoice.payment_mode = PaymentMode.objects.filter(mode__contains = paymode).first()
# 			else:
# 				invoice = InvoiceSells.objects.last()
		
		
# 			invoice.save()
# 			name = product_cell.value
# 			print('name: ' ,name)
# 			import re
# 			item_lst = []
# 			name = name.replace('-',' ')

# 			for name_part in name.split(' '):
# 				print('keyword: ',name_part)
# 				if name_part != None:
# 					for itm in Item.objects.filter(item_name__contains=name_part,price=(sheet.cell(row=ri,column=5).value /  sheet.cell(row=ri,column=4).value)):
# 						exists = False
# 						for ele in item_lst:
# 							if ele['itm'] == itm :
# 								exists = True
# 								ele['occrnc'] += 1
# 						if exists == False:
# 							item_lst.append({'itm':itm,'occrnc':1})
						
# 			print('item_lst : ',item_lst)
# 			max = 0
# 			for ele in item_lst:
# 				if max < ele['occrnc']:
# 					max = ele['occrnc']
		
# 			for ele in item_lst:
# 				if max == ele['occrnc']:
# 					obj = SoldItem(invoice_id=invoice,item=ele['itm'],item_quantity=sheet.cell(row=ri,column=4).value,price=sheet.cell(row=ri,column=5).value)
# 					obj.save()

			# item_lst_dict = {}
			# for obj in item_lst:
			# 	print(item_lst.count(obj))
				# print(obj)


			# soldItm = Item.objects.filter(item_name__contains=([name_part ]),price=  sheet.cell(row=ri,column=4).value /  sheet.cell(row=ri,column=3).value).first()
			# item = SoldItem(invoice_id= invoice,item=soldItm)
			#all([cart_form.is_valid() for cart_form in cart_forms])
			# if 'RAJWADI' in name:
			# Item.objects.get(item_name__ne='RAJWADI')   
			# item.item_quantity = sheet.cell(row=ri,column=3).value
			# item.price = sheet.cell(row=ri,column=4).value
			# item.save()

		# print(cell.value,end=' ')
		# print('')

# readXL()

# import pandas as pd
# # Reading an excel file using Python
# # --------------pandas method-------------
# def pandaReadExcelForMe():
# 	filename = 'KWALITY PRODUCT SALE REPORT 2021 (1).xlsx'
# 	data = pd.read_excel(os.path.join(BASE_DIR,filename),engine='openpyxl')
# 	df = pd.DataFrame(data,columns=['DATE','Total'])
# 	print(df)

# pandaReadExcelForMe()

# Reading an excel file using Python
# --------------xlrd method------------- ********not working************
# import xlrd

# def excelReading(): 
# 	#Give the location of the file
# 	filename = "studentExcelData.xls"
# 	loc = (os.path.join(BASE_DIR,filename))
	
# 	# To open Workbook
# 	wb = xlrd.open_workbook(loc)
# 	sheet = wb.sheet_by_index(0)
	
# 	#For row 0 and column 0
# 	print('cell-val: ',sheet.cell_value(0, 0))

# excelReading()

def home(request):
	sellsForm = InvoiceSellsForm(request.POST or None)
	soldItemForm = SoldItemForm(request.POST or None)
	if request.method=='POST': 
		if sellsForm.is_valid() and soldItemForm.is_valid():
			sellsForm.save()
			print('valid form: ',soldItemForm)
			soldItemForm.save()
			# email = regi_form.cleaned_data.get('email')
			print('valid')
		else:
			# print(regi_form.errors)
			print('errors')
	context = {'three':[1,2,3],'categories':ItemCategory.objects.all(),'items':Item.objects.all(),'sellsForm':sellsForm,'soldItemForm':soldItemForm}
	#item_field = str(context['soldItemForm']['item']).replace('"',"'")
	#context['itemField'] = item_field
	return render(request,'tracker/home.html',context=context)

def fetch_items(request):
	lst = []
	for item in Item.objects.filter(category__id=request.GET.get('categoryId'),current_stock__gt=0):
		lst.append({'id':item.id,'item_name':item.item_name +' ('+ str(item.price) + ')' })
	data={
		'items':lst,
	}
	return JsonResponse(data)

def redirect_page(request):
	#stock_details()
	return redirect('/admin/tracker/invoicesells/add/')

def stock_details():
	# help(pywhatkit)
	# pywhatkit.help
	# pywhatkit.manual()
	
	
	# wait = WebDriverWait(browser, 10000)
	# target = '"temp_group"' #enter contact name here
	# string = "Message by python!" #target msg
	# x_arg = ' //span[contains(@title, ' + target +')]'
	# target = wait.until(ec.presence_of_element_located((By.XPATH, x_arg)))
	# target.click()
	
	# input_box = browser.find_element_by_class_name('_1Plpp')
	# for i in range(100):#loops runs for 100 times
		# input_box.send_keys(string + Keys.ENTER)

	print('coming here')
	title = ''	
	# def end_stock_today():
	lst = [['Date:- '+str(datetime.date.today())],[title]]
	
	sold_items = SoldItem.objects.all()
	if len(sold_items) > 0:
		lst.append(['SELLS'])
		lst.append(['SR\nNO.', 'Name','Qnt.', 'Price'])
		
	sr_no = 0
	for sold_item in sold_items:
		sr_no += 1
		row = [sr_no,sold_item.item.item_name,sold_item.item_quantity,sold_item.price]
		lst.append(row)
	
	# purchased_items = PurchasedItem.objects.all()
	# if len(purchased_items) > 0:
		# lst.append([])
		# lst.append(['PURCHASES'])
		# lst.append(['SR\nNO.', 'Name','Qnt.', 'Price'])
	# # else:
		# # lst.append(['There are no purchases today'])
	# sr_no = 0
	# for purchased_item in purchased_items:
		# sr_no += 1
		# row = [sr_no,purchased_item.item.item_name,purchased_item.item_quantity,purchased_item.price]
		# lst.append(row)
	
	filename = 'KW_Today_Report.pdf'

	# bhavin
	# objects = Appointment.objects.all()

	# end
	pdf = SimpleDocTemplate(
		filename,
		pagesize=letter,
		title="Kwality Wall's Sells & Purchases"
	)
	style = TableStyle([
		# ('SPAN', (0, 0), (-1, 0)),
		# ('SPAN', (0, -1), (-1, -1)),
		# ('FONTSIZE', (0, 0), (0, 0), 20),

		# ('BACKGROUND', (0, 1), (-1, 1), colors.green),
		# ('TEXTCOLOR', (0, 1), (-1, 1), colors.whitesmoke),

		# ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
		# ('FONTSIZE', (0, 1), (-1, -1), 12),
		# # ('RIGHTPADDING', (0, 1), (-1, -1), 0),
		# ('BOTTOMPADDING', (0, 1), (-1, 1), 7),
		# ('TOPPADDING', (0, -1), (-1, -1), 8),

		# ('BOTTOMPADDING', (0, 0), (0, 0), 35),
		('GRID',(0,2),(-1,-1),0.25,colors.black)
	])
	table = Table(lst)
	table.setStyle(style)
	elems = []
	elems.append(table)

	pdf.build(elems)
		
		# p.drawCentredString(274,800,'Orders')
		# p.showPage()
		# p.save()
		# buffer.seek(0)

		# return FileResponse(buffer,as_attachment=False,filename=filename)
	# return FileResponse(open(filename, 'rb'), as_attachment=False, filename=filename)
	#file = open(filename, 'rb')
	
	file_path = os.path.join(BASE_DIR,filename) #file path
	
	browser = webdriver.Chrome(os.path.join(BASE_DIR,'chromedriver'))
	browser.get('https://web.whatsapp.com/')
	sleep(6)
	name = 'Sahil Xocolat'
	user = browser.find_element_by_xpath('//span[@title= "{}"]'.format(name))
	user.click()
	
	attachment_box = browser.find_element_by_xpath('//div[@title= "Attach"]')
	attachment_box.click()
	
	document_box = browser.find_element_by_xpath(
		'//input[@accept=""]'
	)
	document_box.send_keys(file_path)
	
	sleep(3)
	
	send_button = browser.find_element_by_xpath('//span[@data-icon="send-light"]')
	send_button.click()
	
	# #pywhatkit.sendwhatmsg('+919737524730',file,datetime.datetime.now().hour,datetime.datetime.now().minute+1)
	# return HttpResponse('Done')
	# img = Image.open(os.path.join(BASE_DIR,'tracker/mangoRus.jpg'))
	# print(img)
	# # path where the tesseract module is installed 
	# pytesseract.pytesseract.tesseract_cmd =os.path.join(BASE_DIR,'Tesseract-OCR/tesseract.exe')
	# result = pytesseract.image_to_string(img, config='--psm 11')    
	# # write text in a text file and save it to source path    
	# # with open('abc.txt',mode ='w') as file:      
		 # # file.write(result) 
	# print(result)
	# print(datetime.datetime.now().hour)
	#whatsapp_greetings()
	
	#return render(request,'tracker/stock_detail.html',{'items':Item.objects.all(),'result':'result'})
	
# def whatsapp_greetings():
# 	pywhatkit.sendwhatmsg('',"Greetings from xocolate, Please give us feedback how you find your ice cream/corn provided by kwality wall's",datetime.datetime.now().hour,datetime.datetime.now().minute+1)
# 	return None
	
	
	
	